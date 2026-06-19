import pypdf
import os
import re
import json
import openpyxl

def parse_pdf(file_path):
    """
    Parses a PDF file and extracts its text content.
    """
    if not os.path.exists(file_path):
        return f"Error: File not found at {file_path}"

    try:
        with open(file_path, 'rb') as pdf_file:
            reader = pypdf.PdfReader(pdf_file)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
            return text
    except Exception as e:
        return f"An error occurred: {e}"

def structure_text(text):
    """
    Structures the extracted text into a list of atomic components.
    """
    components = []
    current_pillar = ""
    current_section = ""

    # Regex to identify pillars, sections, and actions
    pillar_regex = re.compile(r"Pillar (I|II|III): (.*)")
    # A simple heuristic for sections: title-cased lines that are not actions.
    # And actions are bullet points.
    action_regex = re.compile(r"•\s+(.*)")

    lines = text.split('\n')
    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        pillar_match = pillar_regex.match(line)
        if pillar_match:
            current_pillar = f"Pillar {pillar_match.group(1)}: {pillar_match.group(2)}"
            current_section = "" # Reset section on new pillar
            continue

        action_match = action_regex.match(line)
        if action_match:
            action_text = action_match.group(1).strip()
            # The section is usually the title-cased line(s) before the actions start.
            # Let's look back to find it.
            if not current_section:
                # Look backwards from the current line for a title-cased line
                for j in range(i - 1, -1, -1):
                    prev_line = lines[j].strip()
                    # Heuristic: A section heading is title-cased, not all-caps, and not a pillar.
                    if prev_line and prev_line == prev_line.title() and not prev_line.isupper() and not pillar_regex.match(prev_line):
                         # It's likely a section title. Let's check if the previous line is also part of it.
                        full_section_title = prev_line
                        if j > 0:
                            very_prev_line = lines[j-1].strip()
                            if very_prev_line and very_prev_line == very_prev_line.title() and not very_prev_line.isupper():
                                full_section_title = very_prev_line + " " + full_section_title

                        current_section = full_section_title
                        break
            
            # If we still don't have a section, we can assign a default
            if not current_section:
                current_section = "Uncategorized"


            components.append({
                "source": "America's AI Action Plan July 2025",
                "pillar": current_pillar,
                "section": current_section,
                "action": action_text
            })
        # This is a simple heuristic and might not be perfect.
        # A line is a section if it is title cased and not a pillar or an action
        elif line == line.title() and not line.isupper() and not pillar_regex.match(line) and not action_regex.match(line):
             # This could be a section title. Let's see if the next lines are actions.
             is_section = False
             for k in range(i + 1, min(i + 5, len(lines))):
                 if action_regex.match(lines[k].strip()):
                     is_section = True
                     break
             if is_section:
                current_section = line


    return components

def get_excel_data(file_path, sheet_name=None):
    """
    Reads a specific sheet from an Excel file and returns the data as a list of lists.
    If no sheet_name is provided, it reads the active sheet.
    The first row is assumed to be the header.
    """
    if not os.path.exists(file_path):
        return f"Error: File not found at {file_path}"

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            # openpyxl can return rows with None for all cells for empty rows
            if all(cell is None for cell in row):
                continue
            data.append(list(row))
        return data
    except Exception as e:
        return f"An error occurred while reading the Excel file: {e}"

if __name__ == "__main__":
    excel_path = "Americas-AI-Action-Plan-IVN-Inventory.xlsx"
    excel_data = get_excel_data(excel_path)
    for row in excel_data:
        print(row)

    pdf_path = "C:\\Users\\Basil.White\\OneDrive - USDA\\OCIO-STRATUS Governance Document Working Group - Documents\\Americas-AI-Action-Plan.pdf"
    
    extracted_text = parse_pdf(pdf_path)
    
    if extracted_text.startswith("Error:") or extracted_text.startswith("An error occurred:"):
        print(extracted_text)
    else:
        components = structure_text(extracted_text)
        
        output_file_path = "americas_ai_action_plan_components.json"
        with open(output_file_path, "w", encoding="utf-8") as f:
            json.dump(components, f, indent=2)
        
        print(f"Successfully parsed and saved {len(components)} components to {output_file_path}")
