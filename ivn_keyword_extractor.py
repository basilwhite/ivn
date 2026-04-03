#!/usr/bin/env python3
"""
IVN Keyword Extractor

A Python script that extracts uncommon keywords and key phrases from Excel (.xlsx) files.
Users select an input file, choose which worksheet tab to analyze, and the script
extracts keywords and phrases from the text content.

Usage:
    python ivn_keyword_extractor.py
"""

import re
import json
import csv
from collections import Counter
from pathlib import Path
from typing import List, Dict, Set

# Third-party imports
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

try:
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox
except ImportError:
    print("Error: tkinter is required for file dialogs.")
    exit(1)


# ============================================================================
# STOPWORDS - Common words to filter out during keyword extraction
# ============================================================================
STOPWORDS = {
    'the', 'be', 'to', 'of', 'and', 'a', 'in', 'that', 'have', 'i',
    'it', 'for', 'not', 'on', 'with', 'he', 'as', 'you', 'do', 'at',
    'this', 'but', 'his', 'by', 'from', 'they', 'we', 'say', 'her', 'she',
    'or', 'an', 'will', 'my', 'one', 'all', 'would', 'there', 'their',
    'what', 'so', 'up', 'out', 'if', 'about', 'who', 'get', 'which', 'go',
    'me', 'when', 'make', 'can', 'like', 'time', 'no', 'just', 'him', 'know',
    'take', 'people', 'into', 'year', 'your', 'good', 'some', 'could', 'them',
    'see', 'other', 'than', 'then', 'now', 'look', 'only', 'come', 'its', 'over',
    'think', 'also', 'back', 'after', 'use', 'two', 'how', 'our', 'work', 'first',
    'well', 'way', 'even', 'new', 'want', 'because', 'any', 'these', 'give', 'day',
    'most', 'us', 'is', 'was', 'are', 'been', 'has', 'had', 'were', 'said', 'did',
    'having', 'may', 'should', 'must', 'shall', 'being', 'does', 'such', 'each',
    'through', 'where', 'both', 'those', 'during', 'before', 'herself', 'himself',
    'itself', 'between', 'under', 'above', 'below', 'within', 'without', 'against',    'none', 'more', 'very', 'still', 'here', 'too', 'own', 'same', 'been', 'being',
    'including', 'based', 'using', 'available', 'provides', 'provide', 'requires',
    'addresses', 'supports', 'enable', 'enables', 'meet', 'meets', 'needs',
    'dr', 'it'
}

# Words to exclude from acronym detection (common uppercase words)
ACRONYM_EXCLUDES = {
    'THE', 'AND', 'FOR', 'THIS', 'THAT', 'WITH', 'FROM', 'WILL', 'HAVE',
    'ARE', 'WAS', 'BEEN', 'HAS', 'HAD', 'ALL', 'NOT', 'BUT', 'CAN'
}


def select_xlsx_file() -> str:
    """
    Open a file dialog for the user to select an XLSX file.
    
    Returns:
        str: Path to the selected file, or empty string if cancelled.
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    root.attributes('-topmost', True)  # Bring dialog to front
    
    file_path = filedialog.askopenfilename(
        title="Select Excel File (.xlsx)",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()
    return file_path


def get_workbook_tabs(file_path: str) -> List[str]:
    """
    Read an XLSX file and return a list of all worksheet tab names.
    
    Args:
        file_path: Path to the XLSX file.
        
    Returns:
        List of worksheet names.
    """
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    tab_names = workbook.sheetnames
    workbook.close()
    return tab_names


def select_tab(tab_names: List[str]) -> str:
    """
    Display a dialog for the user to select which tab to process.
    
    Args:
        tab_names: List of available worksheet names.
        
    Returns:
        The selected tab name, or empty string if cancelled.
    """
    selected_tab = ""
    
    def on_select():
        nonlocal selected_tab
        selection = listbox.curselection()
        if selection:
            selected_tab = tab_names[selection[0]]
        root.destroy()
    
    def on_double_click(event):
        on_select()
    
    def on_cancel():
        root.destroy()
    
    root = tk.Tk()
    root.title("Select Worksheet Tab")
    root.geometry("400x300")
    root.attributes('-topmost', True)
    
    # Center the window
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (400 // 2)
    y = (root.winfo_screenheight() // 2) - (300 // 2)
    root.geometry(f"400x300+{x}+{y}")
    
    # Label
    label = tk.Label(root, text="Select a worksheet tab to extract keywords from:", 
                     font=('Arial', 10))
    label.pack(pady=10)
    
    # Listbox with scrollbar
    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
    
    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, font=('Arial', 10),
                         selectmode=tk.SINGLE)
    for tab in tab_names:
        listbox.insert(tk.END, tab)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    listbox.bind('<Double-1>', on_double_click)
    
    scrollbar.config(command=listbox.yview)
    
    # Select first item by default
    if tab_names:
        listbox.selection_set(0)
    
    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)
    
    select_btn = tk.Button(button_frame, text="Select", command=on_select, width=10)
    select_btn.pack(side=tk.LEFT, padx=5)
    
    cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel, width=10)
    cancel_btn.pack(side=tk.LEFT, padx=5)
    
    root.mainloop()
    
    return selected_tab


def normalize_apostrophes(text: str) -> str:
    """
    Replace Unicode apostrophe variants with plain ASCII apostrophe.
    
    This prevents encoding issues where characters like ʼ (modifier letter apostrophe)
    or ' (right single quotation mark) get corrupted in output.
    
    Args:
        text: Input text that may contain Unicode apostrophes.
        
    Returns:
        Text with all apostrophe variants replaced with plain apostrophe (').
    """
    # Unicode apostrophe and quote variants to normalize
    apostrophe_variants = [
        '\u02bc',  # ʼ - Modifier Letter Apostrophe
        '\u2019',  # ' - Right Single Quotation Mark
        '\u2018',  # ' - Left Single Quotation Mark
        '\u0060',  # ` - Grave Accent
        '\u00b4',  # ´ - Acute Accent
        '\u2032',  # ′ - Prime
        '\u2035',  # ‵ - Reversed Prime
    ]
    
    for variant in apostrophe_variants:
        text = text.replace(variant, "'")
    
    return text


def read_xlsx_tab(file_path: str, tab_name: str) -> tuple[List[str], List[str], List[Dict]]:
    """
    Read all text content from a specific worksheet tab.
    
    Args:
        file_path: Path to the XLSX file.
        tab_name: Name of the worksheet to read.
        
    Returns:
        Tuple of (headers, paragraphs, row_data):
        - headers: List of column headers from the first row
        - paragraphs: List of text paragraphs (combined text from each row)
        - row_data: List of dicts containing original column values for each row
    """
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    worksheet = workbook[tab_name]
    
    headers = []
    paragraphs = []
    row_data = []
    
    for idx, row in enumerate(worksheet.iter_rows()):
        # First row contains headers
        if idx == 0:
            headers = [str(cell.value).strip() if cell.value is not None else f"Column{i+1}" 
                      for i, cell in enumerate(row)]
            continue
        
        row_text_parts = []
        row_values = {}
        
        for col_idx, cell in enumerate(row):
            if cell.value is not None:
                cell_text = str(cell.value).strip()
                # Normalize Unicode apostrophes to plain ASCII
                cell_text = normalize_apostrophes(cell_text)
                # Replace hard returns and tabs with single space in Component Description column
                if col_idx < len(headers) and 'Component Description' in headers[col_idx]:
                    cell_text = cell_text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                    # Normalize multiple spaces to single space
                    cell_text = ' '.join(cell_text.split())
                if cell_text:
                    row_text_parts.append(cell_text)
                    # Store original column value
                    if col_idx < len(headers):
                        row_values[headers[col_idx]] = cell_text
        
        # Combine all cells in a row into one paragraph
        if row_text_parts:
            row_text = ' '.join(row_text_parts)
            # Only include paragraphs with 50+ characters
            if len(row_text) >= 50:
                paragraphs.append(row_text)
                row_data.append(row_values)
    
    workbook.close()
    return headers, paragraphs, row_data


def clean_text_for_words(text: str) -> str:
    """
    Clean text for word extraction - convert to lowercase.
    
    Args:
        text: Raw text string.
        
    Returns:
        Cleaned lowercase text string.
    """
    text = text.lower()
    # Remove special characters but keep hyphens, spaces, and numbers
    text = re.sub(r'[^\w\s-]', ' ', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def tokenize(text: str) -> List[str]:
    """
    Split text into words, filtering out short words (3 or fewer characters).
    
    Args:
        text: Cleaned text string.
        
    Returns:
        List of words.
    """
    words = text.split()
    # Filter out words with 3 or fewer characters
    return [word for word in words if len(word) > 3]


def extract_acronyms(text: str) -> List[str]:
    """
    Extract acronyms (all uppercase words 2+ chars) from original text.
    
    Args:
        text: Original text (not lowercased).
        
    Returns:
        List of unique acronyms found.
    """
    # Find all uppercase words (2+ chars) 
    acronym_pattern = r'\b[A-Z][A-Z0-9]{1,}(?:-[A-Z0-9]+)*\b'
    found = re.findall(acronym_pattern, text)
    
    # Filter out excluded common words and dedupe
    acronyms = []
    seen = set()
    for word in found:
        if word not in ACRONYM_EXCLUDES and word not in seen:
            acronyms.append(word)
            seen.add(word)
    
    return acronyms


def extract_capitalized_phrases(text: str) -> List[str]:
    """
    Extract meaningful capitalized phrases (proper nouns, titles, technical terms).
    
    Args:
        text: Original text (not lowercased).
        
    Returns:
        List of capitalized phrases found.
    """
    phrases = []
    
    # Pattern for capitalized multi-word phrases (2-5 words)
    # Matches: "Cloud Computing Policy", "Federal Information Technology"
    cap_phrase_pattern = r'\b([A-Z][a-z]+(?:\s+(?:[A-Z][a-z]+|[A-Z]+|of|and|the|for))+)\b'
    found_phrases = re.findall(cap_phrase_pattern, text)
    
    for phrase in found_phrases:
        # Clean up and validate
        phrase = phrase.strip()
        words = phrase.split()
        # Must have at least 2 words and start with capital
        if len(words) >= 2 and words[0][0].isupper():
            # Skip if it's just stopwords
            non_stop = [w for w in words if w.lower() not in {'of', 'and', 'the', 'for', 'a', 'an'}]
            if len(non_stop) >= 1:
                phrases.append(phrase)
    
    # Also extract patterns with parenthetical acronyms
    # Matches: "Enterprise Cloud Vendor Management (ECVM)"
    paren_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s*\([A-Z]{2,}\)'
    paren_matches = re.findall(paren_pattern, text)
    phrases.extend(paren_matches)
    
    # Dedupe while preserving order
    seen = set()
    unique_phrases = []
    for p in phrases:
        p_lower = p.lower()
        if p_lower not in seen:
            unique_phrases.append(p)
            seen.add(p_lower)
    
    return unique_phrases


def extract_uncommon_words(paragraphs: List[str]) -> Set[str]:
    """
    Extract uncommon words that appear rarely across all paragraphs.
    
    Uncommon words are defined as words that:
    - Appear in only 1-2 paragraphs (rare across document)
    - Are longer than 5 characters (likely to be meaningful)
    - Are not stopwords
    - Appear at least once
    
    Args:
        paragraphs: List of all paragraph texts.
        
    Returns:
        Set of uncommon words found across the document.
    """
    # Track which paragraphs each word appears in
    word_paragraph_count: Dict[str, int] = {}
    
    for paragraph in paragraphs:
        cleaned = clean_text_for_words(paragraph)
        words = tokenize(cleaned)
        
        # Use a set to count each word only once per paragraph
        unique_words_in_para = set(words)
        
        for word in unique_words_in_para:
            if word not in STOPWORDS and len(word) > 5:
                word_paragraph_count[word] = word_paragraph_count.get(word, 0) + 1
    
    # Find words that appear in only 1-2 paragraphs (uncommon)
    total_paragraphs = len(paragraphs)
    threshold = min(2, max(1, total_paragraphs // 10))  # Max 2, or 10% of paragraphs
    
    uncommon = {
        word for word, count in word_paragraph_count.items()
        if count <= threshold
    }
    
    return uncommon


def extract_keywords(text: str, uncommon_words: Set[str] = None) -> Dict[str, List[str]]:
    """
    Extract uncommon keywords and key phrases from text.
    
    Preserves:
    - Acronyms in uppercase (USDA, GAO, FITARA)
    - Meaningful capitalized phrases (Cloud Computing, Service Level Agreement)
    - Important keywords by frequency
    
    Args:
        text: Raw paragraph text.
        uncommon_words: Set of uncommon words to prioritize (optional).
        
    Returns:
        Dictionary with 'keywords', 'phrases', and 'acronyms' lists.
    """
    # Extract acronyms from original text (before lowercasing)
    acronyms = extract_acronyms(text)
    
    # Extract capitalized phrases from original text
    cap_phrases = extract_capitalized_phrases(text)
    
    # Clean text for word frequency analysis
    cleaned = clean_text_for_words(text)
    words = tokenize(cleaned)
    
    # Count word frequencies (excluding stopwords)
    word_counts: Counter[str] = Counter()
    for word in words:
        if word not in STOPWORDS:
            word_counts[word] += 1
    
    # Get top keywords by frequency
    # Prioritize acronyms (add them first), then add frequent words
    keyword_set: Set[str] = set()
    keywords: List[str] = []
    
    # Add acronyms first (lowercased for consistency in keywords)
    for acr in acronyms[:3]:  # Max 3 acronyms in keywords
        acr_lower = acr.lower()
        if acr_lower not in keyword_set:
            keywords.append(acr_lower)
            keyword_set.add(acr_lower)
      # Add most common words
    for word, _ in word_counts.most_common(10):
        if word not in keyword_set and len(keywords) < 7:
            keywords.append(word)
            keyword_set.add(word)
    
    # Add uncommon words that appear in this paragraph
    if uncommon_words:
        paragraph_words = set(words)  # Words from current paragraph
        for word in paragraph_words:
            if (word in uncommon_words and 
                word not in keyword_set and 
                len(keywords) < 10):
                keywords.append(word)
                keyword_set.add(word)
    
    # Build phrases list
    # Start with meaningful capitalized phrases
    phrases: List[str] = []
    phrase_set: Set[str] = set()
    
    for phrase in cap_phrases[:5]:
        p_lower = phrase.lower()
        if p_lower not in phrase_set:
            phrases.append(phrase)
            phrase_set.add(p_lower)
    
    # Add n-gram phrases from lowercased text
    # Extract bigrams (2-word phrases)
    bigram_counts: Counter[str] = Counter()
    for i in range(len(words) - 1):
        if words[i] not in STOPWORDS or words[i + 1] not in STOPWORDS:
            bigram = f"{words[i]} {words[i + 1]}"
            bigram_counts[bigram] += 1
    
    # Extract trigrams (3-word phrases)  
    trigram_counts: Counter[str] = Counter()
    for i in range(len(words) - 2):
        # At least one word should not be a stopword
        if (words[i] not in STOPWORDS or 
            words[i + 1] not in STOPWORDS or 
            words[i + 2] not in STOPWORDS):
            trigram = f"{words[i]} {words[i + 1]} {words[i + 2]}"
            trigram_counts[trigram] += 1
    
    # Add top n-grams that appear more than once
    for phrase, count in trigram_counts.most_common():
        if count > 1 and phrase not in phrase_set and len(phrases) < 7:
            phrases.append(phrase)
            phrase_set.add(phrase)
    
    for phrase, count in bigram_counts.most_common():
        if count > 1 and phrase not in phrase_set and len(phrases) < 7:
            phrases.append(phrase)
            phrase_set.add(phrase)
    
    return {
        'keywords': keywords,
        'phrases': phrases,
        'acronyms': acronyms
    }


def process_paragraphs(paragraphs: List[str], row_data: List[Dict] = None) -> List[Dict]:
    """
    Process all paragraphs and extract keywords from each.
    
    Args:
        paragraphs: List of paragraph texts.
        row_data: Optional list of dicts containing original column values.
        
    Returns:
        List of result dictionaries.
    """
    results = []
    
    # Extract uncommon words from all paragraphs combined
    uncommon_words = extract_uncommon_words(paragraphs)
    
    for index, paragraph in enumerate(paragraphs):
        extracted = extract_keywords(paragraph, uncommon_words)
        result = {
            'rowNumber': index + 1,
            'text': paragraph,
            'keywords': extracted['keywords'],
            'phrases': extracted['phrases'],
            'acronyms': extracted['acronyms']
        }
        
        # Add original column data if available
        if row_data and index < len(row_data):
            result['original_columns'] = row_data[index]
        
        results.append(result)
    
    return results


def display_results(results: List[Dict]) -> None:
    """
    Display extraction results in the console.
    
    Args:
        results: List of result dictionaries.
    """
    print("\n" + "=" * 80)
    print(f"KEYWORD EXTRACTION RESULTS - {len(results)} paragraphs processed")
    print("=" * 80)
    
    for result in results:
        print(f"\n--- Paragraph {result['rowNumber']} ---")
        
        # Truncate text to 300 characters for display
        text_preview = result['text'][:300]
        if len(result['text']) > 300:
            text_preview += "..."
        print(f"Text: {text_preview}")
        
        print(f"Keywords: {', '.join(result['keywords']) if result['keywords'] else '(none)'}")
        print(f"Phrases: {', '.join(result['phrases']) if result['phrases'] else '(none)'}")
        print(f"Acronyms: {', '.join(result['acronyms']) if result['acronyms'] else '(none)'}")
    
    print("\n" + "=" * 80)


def export_csv(results: List[Dict], output_path: str, headers: List[str] = None) -> None:
    """
    Export results to a CSV file.
    
    Args:
        results: List of result dictionaries.
        output_path: Path for the output CSV file.
        headers: Optional list of original column headers from Excel file.
    """
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
        
        # Build header row
        if headers and len(results) > 0 and 'original_columns' in results[0]:
            # Include original columns plus Keywords and Key Phrases
            header_row = ['Row'] + headers + ['Keywords', 'Key Phrases']
        else:
            # Fallback to simple format
            header_row = ['Row', 'Paragraph Text', 'Keywords', 'Key Phrases']
        
        writer.writerow(header_row)
        
        # Write data rows
        for result in results:
            # Combine acronyms with keywords for output
            all_keywords = result['acronyms'] + result['keywords']
            # Dedupe while preserving order
            seen = set()
            unique_keywords = []
            for kw in all_keywords:
                kw_lower = kw.lower()
                if kw_lower not in seen:
                    unique_keywords.append(kw)
                    seen.add(kw_lower)
            
            # Build row data
            if 'original_columns' in result and headers:
                # Include original column values
                row_data = [result['rowNumber']]
                for header in headers:
                    row_data.append(result['original_columns'].get(header, ''))
                row_data.extend([
                    ', '.join(unique_keywords),
                    ', '.join(result['phrases'])
                ])
            else:
                # Fallback to simple format
                row_data = [
                    result['rowNumber'],
                    result['text'],
                    ', '.join(unique_keywords),
                    ', '.join(result['phrases'])
                ]
            
            writer.writerow(row_data)
    
    print(f"CSV exported to: {output_path}")


def export_json(results: List[Dict], output_path: str) -> None:
    """
    Export results to a JSON file.
    
    Args:
        results: List of result dictionaries.
        output_path: Path for the output JSON file.
    """
    with open(output_path, 'w', encoding='utf-8') as jsonfile:
        json.dump(results, jsonfile, indent=2, ensure_ascii=False)
    
    print(f"JSON exported to: {output_path}")


def main():
    """Main entry point for the keyword extractor."""
    print("\n" + "=" * 80)
    print("IVN KEYWORD EXTRACTOR")
    print("=" * 80)
    
    # Step 1: Prompt user to select XLSX file
    print("\nStep 1: Select an Excel (.xlsx) file...")
    file_path = select_xlsx_file()
    
    if not file_path:
        print("No file selected. Exiting.")
        return
    
    print(f"Selected file: {file_path}")
    
    # Validate file extension
    if not file_path.lower().endswith('.xlsx'):
        print("Error: Please select a valid .xlsx file.")
        return
    
    # Step 2: Get list of tabs in the workbook
    print("\nStep 2: Reading workbook tabs...")
    try:
        tab_names = get_workbook_tabs(file_path)
    except Exception as e:
        print(f"Error reading workbook: {e}")
        return
    
    if not tab_names:
        print("Error: No worksheets found in the workbook.")
        return
    
    print(f"Found {len(tab_names)} tab(s): {', '.join(tab_names)}")
    
    # Step 3: Prompt user to select which tab to use
    print("\nStep 3: Select a worksheet tab...")
    selected_tab = select_tab(tab_names)
    
    if not selected_tab:
        print("No tab selected. Exiting.")
        return
    
    print(f"Selected tab: {selected_tab}")
    
    # Step 4: Read content from selected tab
    print("\nStep 4: Extracting text from worksheet...")
    try:
        headers, paragraphs, row_data = read_xlsx_tab(file_path, selected_tab)
    except Exception as e:
        print(f"Error reading worksheet: {e}")
        return
    
    if not paragraphs:
        print("No paragraphs found (minimum 50 characters per row).")
        return
    
    print(f"Found {len(paragraphs)} paragraph(s) to process.")
    
    # Step 5: Extract keywords from each paragraph
    print("\nStep 5: Extracting keywords and phrases...")
    results = process_paragraphs(paragraphs, row_data)
    
    # Step 6: Display results
    display_results(results)
    
    # Step 7: Export results
    base_path = Path(file_path).parent
    base_name = Path(file_path).stem
    
    csv_path = base_path / f"{base_name}_keywords.csv"
    json_path = base_path / f"{base_name}_keywords.json"
    
    print("\nStep 6: Exporting results...")
    export_csv(results, str(csv_path), headers)
    export_json(results, str(json_path))
    
    print("\n" + "=" * 80)
    print("EXTRACTION COMPLETE!")
    print("=" * 80)
    
    # Show completion message
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(
        "Extraction Complete",
        f"Successfully extracted keywords from {len(results)} paragraphs.\n\n"
        f"CSV saved to:\n{csv_path}\n\n"
        f"JSON saved to:\n{json_path}"
    )
    root.destroy()


if __name__ == "__main__":
    main()
