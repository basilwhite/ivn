# Filename: extract_citations.py


import requests
import PyPDF2
import re
import os
import tempfile
import time
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def sanitize_text(text):
    return re.sub(r"[\r\n]+", " ", text).strip()


def clean_citation(citation):
    # Normalize U.S. Code references
    citation = re.sub(
        r"\b(\d+)\s*(U\.S\.C\.|USC|U\.S\. Code)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9()]*)?)",
        r"\1 USC \3",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize CFR references
    citation = re.sub(
        r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9()]*)?)",
        r"\1 CFR \3",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Executive Order references
    citation = re.sub(
        r"\b(E\.O\.|Executive\s*Order)\s*(\d+)",
        r"Executive Order \2",
        citation,
        flags=re.IGNORECASE
    )


    citation = re.sub(
        r"\bEO\s+(\d+)\b",
        r"Executive Order \1",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Public Laws
    citation = re.sub(
        r"\b(Public\s+Law|P\.L\.)\s*(\d{1,3}[-–]\d{1,4})",
        r"Public Law \2",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Acts
    citation = re.sub(
        r"\bAct\s+of\s+(\d{4})",
        r"Act of \1",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Title references
    citation = re.sub(
        r"\bTitle\s+(\d+)",
        r"Title \1",
        citation,
        flags=re.IGNORECASE
    )


    return citation


def get_browser_headers():
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/pdf",
        "Connection": "keep-alive"
    }


def download_pdf(url):
    try:
        session = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=5,
            status_forcelist=[500, 502, 503, 504],
            raise_on_status=False,
        )
        session.mount('https://', HTTPAdapter(max_retries=retries))


        response = session.get(url, headers=get_browser_headers(), stream=True, timeout=60)
        response.raise_for_status()


        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()


        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        with open("failed_downloads.txt", "a") as f:
            f.write(url + "\n")
        return None


def extract_toc(reader):
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc


def infer_section_name(toc, page_num, context, page_text):
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:
                return section
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"


def extract_us_code_citations(pdf_path, url):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)
            citations = []


            citation_pattern = (
                r"\b(\d+)\s*(U\.S\.C\.|USC|U\.S\. Code)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"(E\.O\.|Executive\s*Order)\s*(\d+)|"
                r"\bEO\s+(\d+)\b|"
                r"\bPublic\s+Law\s+\d{1,3}[-–]\d{1,4}\b|"
                r"\bP\.L\.\s*\d{1,3}[-–]\d{1,4}\b|"
                r"\bAct\s+of\s+\d{4}\b|"
                r"\bTitle\s+\d+\b"
            )


            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                if text:
                    matches = re.finditer(citation_pattern, text, re.IGNORECASE)
                    for match in matches:
                        citation_text = match.group(0)
                        citation = clean_citation(citation_text)
                        start, end = match.start(), match.end()
                        context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                        section_name = infer_section_name(toc, page_num + 1, context, text)
                        citation_page_url = f"{url}#page={page_num + 1}"
                        citations.append((citation, citation_page_url, section_name, context, url))
        return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []


def process_url(url):
    temp_file = download_pdf(url)
    if not temp_file:
        return []
    try:
        return extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)


def save_to_excel(data, filename="extracted_citations.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")


def main():
    url_list = [        
        "https://www.usda.gov/sites/default/files/documents/usda-fy-2024-annual-performance-report.pdf",
        "https://www.usda.gov/sites/default/files/documents/usda-fy-2024-annual-performance-report.pdf",
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))
        time.sleep(3)  # pause between downloads to mimic human browsing


    save_to_excel(all_citations)


if __name__ == "__main__":
    main()






