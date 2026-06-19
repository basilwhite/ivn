#!/usr/bin/env python3
"""
IVN Keyword Crosswalk Analysis Tool
====================================
This script analyzes word frequencies in governance document components to identify
uncommon words for cross-referencing between different document sets.
"""

import pandas as pd
import numpy as np
import re
import string
import os
import sys
import warnings
from collections import Counter
import nltk
from nltk.corpus import brown
from openpyxl import load_workbook
from datetime import datetime

warnings.filterwarnings('ignore')

class IVNKeywordCrosswalk:
    """Main class for IVN Keyword Crosswalk analysis."""
    
    def __init__(self, input_file='ivntest.xlsx', output_file=None):
        """Initialize the analyzer with input and output file paths."""
        self.input_file = input_file
        if output_file is None:
            timestamp = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            self.output_file = f'ivn-keyword-crosswalk-{timestamp}.xlsx'
        else:
            self.output_file = output_file
        self.common_words_data = None
        self.strings_to_ignore = None
        self.ivn_word_frequencies = None
        self.uncommon_scores = None
        
    def normalize_text(self, text):
        """
        Normalize text according to specified rules:
        1. Convert to lowercase
        2. Remove all punctuation except hyphens
        3. Split on whitespace
        """
        if not isinstance(text, str):
            return []
        
        # Convert to lowercase
        text = text.lower()
        
        # Keep hyphens for hyphenated words, remove other punctuation
        # First, replace underscores and other word connectors with space
        text = re.sub(r'[_\t\n\r\f\v]', ' ', text)
        
        # Remove all punctuation except hyphens
        # This regex keeps letters, numbers, hyphens, and whitespace
        text = re.sub(r'[^\w\s-]', '', text)
        
        # Replace multiple spaces with single space
        text = re.sub(r'\s+', ' ', text).strip()
        
        # Split on whitespace
        words = text.split()
        
        return words
    
    def get_common_words_data(self):
        """
        Get common English words data using tiered approach:
        1. Try to download and use NLTK Brown corpus
        2. Fallback to bundled COCA-based list
        3. Exit with instructions if all fails
        """
        print("Loading common English words data...")
        
        # Try NLTK Brown corpus first
        try:
            nltk.data.find('corpora/brown')
            print("Using NLTK Brown corpus...")
            
            # Get words from Brown corpus
            words = brown.words()
            total_words = len(words)
            
            # Count frequencies
            word_counts = Counter(words)
            
            # Convert to ratios
            common_words = {}
            for word, count in word_counts.items():
                # Normalize word (same as our text processing)
                normalized_words = self.normalize_text(word)
                for norm_word in normalized_words:
                    if norm_word:  # Skip empty strings
                        common_words[norm_word] = common_words.get(norm_word, 0) + count
            
            # Recalculate total based on normalized words
            total_normalized = sum(common_words.values())
            
            # Convert to ratios
            common_word_ratios = {}
            for word, count in common_words.items():
                common_word_ratios[word] = count / total_normalized
            
            # Get top 100 words for strings-to-ignore
            top_100_words = [word for word, _ in Counter(words).most_common(100)]
            top_100_normalized = []
            for word in top_100_words:
                normalized = self.normalize_text(word)
                top_100_normalized.extend(normalized)
            
            # Get unique normalized top 100 words
            strings_to_ignore = list(set(top_100_normalized))[:100]
            
            self.common_words_data = common_word_ratios
            self.strings_to_ignore = strings_to_ignore
            
            return True
            
        except (LookupError, OSError) as e:
            print(f"NLTK Brown corpus not available: {e}")
            print("Falling back to bundled COCA-based list...")
            
            # Fallback to bundled COCA-based list
            return self._use_bundled_coca_list()
    
    def _use_bundled_coca_list(self):
        """Use bundled COCA-based list of common English words."""
        try:
            # This is a realistic frequency list based on COCA top 10,000 words
            # Format: word, frequency_ratio
            bundled_data = {
                'the': 0.052, 'be': 0.036, 'to': 0.032, 'of': 0.031, 'and': 0.027,
                'a': 0.024, 'in': 0.021, 'that': 0.019, 'have': 0.017, 'i': 0.016,
                'it': 0.015, 'for': 0.014, 'not': 0.013, 'on': 0.012, 'with': 0.012,
                'he': 0.011, 'as': 0.010, 'you': 0.010, 'do': 0.009, 'at': 0.009,
                'this': 0.009, 'but': 0.008, 'his': 0.008, 'by': 0.008, 'from': 0.008,
                'they': 0.007, 'we': 0.007, 'say': 0.007, 'her': 0.007, 'she': 0.007,
                'or': 0.007, 'an': 0.006, 'will': 0.006, 'my': 0.006, 'one': 0.006,
                'all': 0.006, 'would': 0.006, 'there': 0.006, 'their': 0.006, 'what': 0.006,
                'so': 0.005, 'up': 0.005, 'out': 0.005, 'if': 0.005, 'about': 0.005,
                'who': 0.005, 'get': 0.005, 'which': 0.005, 'go': 0.005, 'me': 0.005,
                'when': 0.005, 'make': 0.005, 'can': 0.004, 'like': 0.004, 'time': 0.004,
                'no': 0.004, 'just': 0.004, 'him': 0.004, 'know': 0.004, 'take': 0.004,
                'people': 0.004, 'into': 0.004, 'year': 0.004, 'your': 0.004, 'good': 0.004,
                'some': 0.004, 'could': 0.004, 'them': 0.004, 'see': 0.004, 'other': 0.004,
                'than': 0.004, 'then': 0.004, 'now': 0.004, 'look': 0.004, 'only': 0.004,
                'come': 0.004, 'its': 0.004, 'over': 0.004, 'think': 0.004, 'also': 0.004,
                'back': 0.004, 'after': 0.003, 'use': 0.003, 'two': 0.003, 'how': 0.003,
                'our': 0.003, 'work': 0.003, 'first': 0.003, 'well': 0.003, 'way': 0.003,
                'even': 0.003, 'new': 0.003, 'want': 0.003, 'because': 0.003, 'any': 0.003,
                'these': 0.003, 'give': 0.003, 'day': 0.003, 'most': 0.003, 'us': 0.003
            }
            
            # Extend with more common words to reach ~10,000 (simplified for example)
            # In production, this would be a complete COCA 10,000 word list
            additional_words = [
                'system', 'data', 'information', 'process', 'control',
                'management', 'security', 'requirements', 'design', 'development',
                'implementation', 'testing', 'maintenance', 'operations', 'support',
                'quality', 'risk', 'compliance', 'governance', 'framework',
                'standard', 'policy', 'procedure', 'guideline', 'documentation',
                'verification', 'validation', 'assessment', 'audit', 'monitoring',
                'reporting', 'analysis', 'evaluation', 'review', 'approval',
                'authorization', 'certification', 'accreditation', 'integration',
                'deployment', 'configuration', 'administration', 'supervision',
                'coordination', 'collaboration', 'communication', 'training',
                'education', 'awareness', 'preparedness', 'readiness', 'response',
                'recovery', 'continuity', 'availability', 'confidentiality', 'integrity',
                'authentication', 'authorization', 'accountability', 'traceability',
                'reliability', 'safety', 'privacy', 'protection', 'defense', 'resilience'
            ]
            
            # Add additional words with lower frequencies
            base_freq = 0.0001
            for i, word in enumerate(additional_words):
                bundled_data[word] = base_freq * (0.9 ** i)
            
            self.common_words_data = bundled_data
            
            # Create strings-to-ignore from top 100 words in bundled data
            sorted_words = sorted(bundled_data.items(), key=lambda x: x[1], reverse=True)
            self.strings_to_ignore = [word for word, _ in sorted_words[:100]]
            
            print("Using bundled COCA-based list (simplified for example)")
            return True
            
        except Exception as e:
            print(f"Failed to use bundled list: {e}")
            print("\n" + "="*60)
            print("SETUP INSTRUCTIONS:")
            print("="*60)
            print("1. Install NLTK data by running:")
            print("   python -m nltk.downloader brown")
            print("\n2. Or ensure the bundled word list is properly configured.")
            print("\n3. Run the script again after completing setup.")
            print("="*60)
            return False
    
    def load_input_data(self):
        """Load and validate input Excel file."""
        print(f"Loading input file: {self.input_file}")
        
        try:
            # Read the Excel file
            excel_file = pd.ExcelFile(self.input_file)
            
            # Check required tabs (flexible matching)
            required_tabs = ['Components', 'to-be-crosswalked']
            tab_mapping = {}
            
            for req_tab in required_tabs:
                # Create normalized version for matching
                req_normalized = req_tab.lower().replace('-', '').replace(' ', '')
                
                found_tab = None
                for sheet in excel_file.sheet_names:
                    sheet_normalized = sheet.lower().replace('-', '').replace(' ', '')
                    if req_normalized == sheet_normalized:
                        found_tab = sheet
                        break
                
                if found_tab:
                    tab_mapping[req_tab] = found_tab
                else:
                    print(f"ERROR: Required tab '{req_tab}' not found (normalized search)")
                    print(f"Available tabs: {excel_file.sheet_names}")
                    sys.exit(1)
            
            # Load the tabs using the correct case
            self.components_df = excel_file.parse(tab_mapping['Components'])
            self.tb_crosswalked_df = excel_file.parse(tab_mapping['to-be-crosswalked'])
            
            # Check for required columns (flexible matching)
            required_col = 'Component Description'
            req_col_normalized = required_col.lower().replace(' ', '').replace('_', '')
            
            comp_desc_col = None
            for col in self.components_df.columns:
                col_normalized = col.lower().replace(' ', '').replace('_', '')
                if col_normalized == req_col_normalized:
                    comp_desc_col = col
                    break
            
            tb_desc_col = None
            for col in self.tb_crosswalked_df.columns:
                col_normalized = col.lower().replace(' ', '').replace('_', '')
                if col_normalized == req_col_normalized:
                    tb_desc_col = col
                    break
            
            if not comp_desc_col:
                print(f"ERROR: '{required_col}' column missing in Components tab")
                print(f"Available columns: {list(self.components_df.columns)}")
                sys.exit(1)
            
            if not tb_desc_col:
                print(f"ERROR: '{required_col}' column missing in to-be-crosswalked tab")
                print(f"Available columns: {list(self.tb_crosswalked_df.columns)}")
                sys.exit(1)
            
            # Rename columns to standard names for consistency
            self.components_df = self.components_df.rename(columns={comp_desc_col: required_col})
            self.tb_crosswalked_df = self.tb_crosswalked_df.rename(columns={tb_desc_col: required_col})
            
            print(f"Loaded {len(self.components_df)} rows from Components tab")
            print(f"Loaded {len(self.tb_crosswalked_df)} rows from to-be-crosswalked tab")
            
            return True
            
        except FileNotFoundError:
            print(f"ERROR: Input file not found: {self.input_file}")
            sys.exit(1)
        except Exception as e:
            print(f"ERROR loading input file: {e}")
            sys.exit(1)
    
    def analyze_ivn_word_frequencies(self):
        """Analyze word frequencies in IVN component descriptions."""
        print("Analyzing IVN word frequencies...")
        
        # Combine all descriptions
        all_descriptions = pd.concat([
            self.components_df['Component Description'].dropna(),
            self.tb_crosswalked_df['Component Description'].dropna()
        ])
        
        # Process all words
        all_words = []
        for desc in all_descriptions:
            words = self.normalize_text(str(desc))
            all_words.extend(words)
        
        # Calculate frequencies
        word_counts = Counter(all_words)
        total_words = len(all_words)
        
        # Convert to ratios
        ivn_word_ratios = {}
        for word, count in word_counts.items():
            ivn_word_ratios[word] = count / total_words
        
        self.ivn_word_frequencies = ivn_word_ratios
        print(f"Analyzed {total_words} words, {len(word_counts)} unique words")
        
        return ivn_word_ratios
    
    def calculate_uncommon_scores(self):
        """Calculate uncommonality scores for IVN words."""
        print("Calculating uncommonality scores...")
        
        uncommon_scores = {}
        
        for word, ivn_ratio in self.ivn_word_frequencies.items():
            # Get common English ratio (use small epsilon if not found)
            common_ratio = self.common_words_data.get(word, 1e-10)
            
            # Calculate uncommonality score
            if common_ratio > 0:
                uncommon_score = ivn_ratio / common_ratio
            else:
                uncommon_score = 1e6  # Very high score for words not in common list
            
            uncommon_scores[word] = uncommon_score
        
        self.uncommon_scores = uncommon_scores
        return uncommon_scores
    
    def get_atypical_keywords(self, description):
        """Extract atypical keywords from a description."""
        if not isinstance(description, str):
            return ""
        
        # Normalize and get words
        words = self.normalize_text(description)
        
        # Filter out strings to ignore
        filtered_words = [word for word in words if word not in self.strings_to_ignore]
        
        # Sort by uncommonality score (descending)
        sorted_words = sorted(
            filtered_words,
            key=lambda x: self.uncommon_scores.get(x, 0),
            reverse=True
        )
        
        # Remove duplicates while preserving order
        unique_words = []
        seen = set()
        for word in sorted_words:
            if word not in seen:
                seen.add(word)
                unique_words.append(word)
        
        return ', '.join(unique_words)
    
    def process_dataframes(self):
        """Add atypical keywords to both dataframes."""
        print("Processing dataframes to add atypical keywords...")
        
        # Process Components dataframe
        self.components_df['atypical-keywords'] = self.components_df['Component Description'].apply(
            self.get_atypical_keywords
        )
        
        # Process to-be-crosswalked dataframe
        self.tb_crosswalked_df['atypical-keywords'] = self.tb_crosswalked_df['Component Description'].apply(
            self.get_atypical_keywords
        )
    
    def check_crosswalk_size(self):
        """Check the total number of potential crosswalk rows and warn if too large."""
        total_rows = len(self.components_df) * len(self.tb_crosswalked_df)
        excel_max_rows = 1048576
        
        print(f"Components tab: {len(self.components_df)} rows")
        print(f"To-be-crosswalked tab: {len(self.tb_crosswalked_df)} rows")
        print(f"Potential crosswalk pairs: {total_rows:,} rows")
        
        if total_rows > excel_max_rows:
            print(f"WARNING: {total_rows:,} rows exceeds Excel's maximum of {excel_max_rows:,} rows!")
            print("Consider reducing the dataset size or increasing the similarity threshold.")
            return False
        
        return True
    
    def sample_similarity_scores(self, sample_size=1000):
        """Sample similarity scores from potential pairs to determine recommended threshold."""
        print(f"Sampling {sample_size} similarity scores to determine recommended threshold...")
        
        # Add prefixes to column names
        dataset_df = self.components_df.copy()
        dataset_df.columns = [f'dataset_{col}' for col in dataset_df.columns]
        
        new_component_df = self.tb_crosswalked_df.copy()
        new_component_df.columns = [f'new_component_{col}' for col in new_component_df.columns]
        
        similarity_scores = []
        count = 0
        
        # Sample from the first dataset rows and all new component rows
        for _, dataset_row in dataset_df.iterrows():
            for _, new_comp_row in new_component_df.iterrows():
                # Calculate similarity
                dataset_keywords = set(dataset_row['dataset_atypical-keywords'].split(', ')) if dataset_row['dataset_atypical-keywords'] else set()
                new_comp_keywords = set(new_comp_row['new_component_atypical-keywords'].split(', ')) if new_comp_row['new_component_atypical-keywords'] else set()
                
                intersection = len(dataset_keywords & new_comp_keywords)
                union = len(dataset_keywords | new_comp_keywords)
                similarity_score = intersection / union if union > 0 else 0
                
                similarity_scores.append(similarity_score)
                count += 1
                
                if count >= sample_size:
                    break
            if count >= sample_size:
                break
        
        if similarity_scores:
            avg_similarity = sum(similarity_scores) / len(similarity_scores)
            print(f"Average similarity score from sample: {avg_similarity:.4f}")
            return avg_similarity
        else:
            print("No similarity scores sampled")
            return 0.0
    
    def get_similarity_threshold(self):
        """Prompt user for similarity threshold with recommended default."""
        recommended_threshold = self.sample_similarity_scores()
        
        while True:
            try:
                user_input = input(f"Enter minimum similarity threshold (recommended: {recommended_threshold:.4f}, press Enter to accept): ").strip()
                
                if user_input == "":
                    threshold = recommended_threshold
                    print(f"Using recommended threshold: {threshold:.4f}")
                else:
                    threshold = float(user_input)
                    if 0 <= threshold <= 1:
                        print(f"Using custom threshold: {threshold:.4f}")
                    else:
                        print("Threshold must be between 0 and 1. Please try again.")
                        continue
                
                return threshold
                
            except ValueError:
                print("Invalid input. Please enter a number between 0 and 1, or press Enter for recommended value.")
    
    def create_crosswalk_dataframe(self, similarity_threshold=0.0):
        """Create the keywords-crosswalk dataframe with similarity scores above threshold."""
        print(f"Creating keywords-crosswalk dataframe (threshold: {similarity_threshold:.4f})...")
        
        # Add prefixes to column names
        dataset_df = self.components_df.copy()
        dataset_df.columns = [f'dataset_{col}' for col in dataset_df.columns]
        
        new_component_df = self.tb_crosswalked_df.copy()
        new_component_df.columns = [f'new_component_{col}' for col in new_component_df.columns]
        
        # Calculate total iterations
        total_iterations = len(dataset_df) * len(new_component_df)
        print(f"Processing {total_iterations:,} potential pairs...")
        
        # Create Cartesian product with threshold filtering
        crosswalk_data = []
        processed_count = 0
        start_time = datetime.now()
        last_update_time = start_time
        
        for i, (_, dataset_row) in enumerate(dataset_df.iterrows()):
            for j, (_, new_comp_row) in enumerate(new_component_df.iterrows()):
                # Calculate similarity based on atypical-keywords
                dataset_keywords = set(dataset_row['dataset_atypical-keywords'].split(', ')) if dataset_row['dataset_atypical-keywords'] else set()
                new_comp_keywords = set(new_comp_row['new_component_atypical-keywords'].split(', ')) if new_comp_row['new_component_atypical-keywords'] else set()
                
                # Jaccard similarity
                intersection = len(dataset_keywords & new_comp_keywords)
                union = len(dataset_keywords | new_comp_keywords)
                similarity_score = intersection / union if union > 0 else 0
                
                # Only include if above threshold
                if similarity_score >= similarity_threshold:
                    # Create combined row
                    combined_row = {}
                    combined_row.update(dataset_row)
                    combined_row.update(new_comp_row)
                    combined_row['similarity_score'] = similarity_score
                    
                    # Truncate keyword strings to prevent Excel cell limit issues (32,767 chars)
                    EXCEL_CELL_LIMIT = 32000  # Leave some buffer
                    
                    shared = ', '.join(sorted(dataset_keywords & new_comp_keywords))
                    dataset_unique = ', '.join(sorted(dataset_keywords - new_comp_keywords))
                    new_comp_unique = ', '.join(sorted(new_comp_keywords - dataset_keywords))
                    
                    combined_row['shared_keywords'] = shared[:EXCEL_CELL_LIMIT] + ('...' if len(shared) > EXCEL_CELL_LIMIT else '')
                    combined_row['dataset_unique_keywords'] = dataset_unique[:EXCEL_CELL_LIMIT] + ('...' if len(dataset_unique) > EXCEL_CELL_LIMIT else '')
                    combined_row['new_component_unique_keywords'] = new_comp_unique[:EXCEL_CELL_LIMIT] + ('...' if len(new_comp_unique) > EXCEL_CELL_LIMIT else '')
                    
                    crosswalk_data.append(combined_row)
                
                processed_count += 1
                
                # Update progress every 1000 pairs or 10% completion
                current_time = datetime.now()
                time_since_last_update = (current_time - last_update_time).total_seconds()
                
                if processed_count % 1000 == 0 or time_since_last_update >= 10:
                    elapsed_time = (current_time - start_time).total_seconds()
                    pairs_remaining = total_iterations - processed_count
                    
                    if processed_count > 0:
                        avg_time_per_pair = elapsed_time / processed_count
                        estimated_remaining = pairs_remaining * avg_time_per_pair
                        
                        # Format time remaining
                        hours = int(estimated_remaining // 3600)
                        minutes = int((estimated_remaining % 3600) // 60)
                        seconds = int(estimated_remaining % 60)
                        time_remaining_str = f"{hours}:{minutes:02d}:{seconds:02d}"
                        
                        print(f"Processed: {processed_count:,} | Remaining: {pairs_remaining:,} | "
                              f"Time remaining: {time_remaining_str}")
                    
                    last_update_time = current_time
        
        crosswalk_df = pd.DataFrame(crosswalk_data)
        
        # Sort by similarity score descending
        if not crosswalk_df.empty:
            crosswalk_df = crosswalk_df.sort_values('similarity_score', ascending=False)
        
        print(f"Created crosswalk with {len(crosswalk_df)} pairs (above threshold {similarity_threshold:.4f})")
        return crosswalk_df
    
    def create_output_excel(self):
        """Create the output Excel file with all required tabs."""
        print(f"Creating output file: {self.output_file}")
        
        try:
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                # 1. Common-words tab
                common_words_df = pd.DataFrame(
                    list(self.common_words_data.items()),
                    columns=['Word', 'Frequency Ratio']
                ).sort_values('Frequency Ratio', ascending=False)
                common_words_df.to_excel(writer, sheet_name='common-words', index=False)
                
                # 2. Common-ivn-words tab
                common_ivn_words_df = pd.DataFrame(
                    list(self.ivn_word_frequencies.items()),
                    columns=['Word', 'Frequency Ratio']
                ).sort_values('Frequency Ratio', ascending=False)
                common_ivn_words_df.to_excel(writer, sheet_name='common-ivn-words', index=False)
                
                # 3. Strings-to-ignore tab (check if exists)
                if os.path.exists(self.output_file):
                    try:
                        existing_wb = load_workbook(self.output_file)
                        if 'strings-to-ignore' in existing_wb.sheetnames:
                            # Read existing strings-to-ignore
                            existing_df = pd.read_excel(self.output_file, sheet_name='strings-to-ignore')
                            strings_to_ignore_df = existing_df
                            print("Preserved existing strings-to-ignore tab with user additions")
                        else:
                            # Create new strings-to-ignore
                            strings_to_ignore_df = pd.DataFrame(self.strings_to_ignore, columns=['Strings to Ignore'])
                    except:
                        # Create new strings-to-ignore
                        strings_to_ignore_df = pd.DataFrame(self.strings_to_ignore, columns=['Strings to Ignore'])
                else:
                    # Create new strings-to-ignore
                    strings_to_ignore_df = pd.DataFrame(self.strings_to_ignore, columns=['Strings to Ignore'])
                
                strings_to_ignore_df.to_excel(writer, sheet_name='strings-to-ignore', index=False)
                
                # 4. Uncommon-ivn-words tab
                uncommon_data = []
                for word, ivn_ratio in self.ivn_word_frequencies.items():
                    if word not in self.strings_to_ignore:
                        common_ratio = self.common_words_data.get(word, 1e-10)
                        uncommon_score = ivn_ratio / common_ratio if common_ratio > 0 else 1e6
                        uncommon_data.append({
                            'Word': word,
                            'IVN Frequency Ratio': ivn_ratio,
                            'Common Frequency Ratio': common_ratio,
                            'Uncommonality Score': uncommon_score
                        })
                
                uncommon_ivn_words_df = pd.DataFrame(uncommon_data)
                uncommon_ivn_words_df = uncommon_ivn_words_df.sort_values('Uncommonality Score', ascending=False)
                uncommon_ivn_words_df.to_excel(writer, sheet_name='uncommon-ivn-words', index=False)
                
                # 5. Keywords-dataset tab
                self.components_df.to_excel(writer, sheet_name='keywords-dataset', index=False)
                
                # 6. New-components-keywords tab
                self.tb_crosswalked_df.to_excel(writer, sheet_name='new-components-keywords', index=False)
                
                # 7. Keywords-crosswalk tab
                self.crosswalk_df.to_excel(writer, sheet_name='keywords-crosswalk', index=False)
            
            print("Output Excel file created successfully")
            
        except Exception as e:
            print(f"ERROR: Failed to create Excel file: {e}")
            print("This may be due to:")
            print("- Very large dataset exceeding Excel limits")
            print("- Special characters in data causing corruption")
            print("- Insufficient disk space or permissions")
            print("- Memory issues during file creation")
            return False
            
        return True
    
    def run(self):
        """Execute the complete analysis pipeline."""
        print("="*60)
        print("IVN Keyword Crosswalk Analysis")
        print("="*60)
        
        # Step 1: Get common words data
        if not self.get_common_words_data():
            return False
        
        # Step 2: Load input data
        self.load_input_data()
        
        # Step 3: Analyze IVN word frequencies
        self.analyze_ivn_word_frequencies()
        
        # Step 4: Calculate uncommon scores
        self.calculate_uncommon_scores()
        
        # Step 5: Process dataframes
        self.process_dataframes()
        
        # Step 6: Check crosswalk size
        if not self.check_crosswalk_size():
            print("Consider adjusting your dataset or threshold to reduce output size.")
        
        # Step 7: Get similarity threshold
        self.similarity_threshold = self.get_similarity_threshold()
        
        # Step 8: Create crosswalk dataframe
        self.crosswalk_df = self.create_crosswalk_dataframe(self.similarity_threshold)
        
        # Step 9: Create output Excel
        if not self.create_output_excel():
            print("Excel file creation failed. Analysis results are still available in memory.")
            print("Consider using a higher similarity threshold to reduce output size.")
            return False
        
        print("="*60)
        print("Analysis complete!")
        print(f"Output saved to: {self.output_file}")
        print("="*60)
        
        return True


def main():
    """Main function to run the IVN Keyword Crosswalk analysis."""
    analyzer = IVNKeywordCrosswalk()
    analyzer.run()


if __name__ == "__main__":
    main()