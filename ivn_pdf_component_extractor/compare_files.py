import openpyxl

# Load both files
wb1 = openpyxl.load_workbook(r'HR_2.0_memo_12-10-2025_2025-12-18-11-16.xlsx')
wb2 = openpyxl.load_workbook(r'corrected-HR_2.0_memo_12-10-2025_2025-12-18-11-16.xlsx')

ws1 = wb1.active
ws2 = wb2.active

with open('comparison_output.txt', 'w', encoding='utf-8') as f:
    f.write("=" * 60 + "\n")
    f.write("SCRIPT OUTPUT (Column B - Component Name)\n")
    f.write("=" * 60 + "\n")
    for i in range(1, ws1.max_row + 1):
        val = ws1.cell(row=i, column=2).value
        if val and len(str(val)) > 80:
            val = val[:80] + "..."
        f.write(f"Row {i}: {val}\n")

    f.write("\n")
    f.write("=" * 60 + "\n")
    f.write("CORRECTED (Column B - Component Name)\n")
    f.write("=" * 60 + "\n")
    for i in range(1, ws2.max_row + 1):
        val = ws2.cell(row=i, column=2).value
        if val and len(str(val)) > 80:
            val = val[:80] + "..."
        f.write(f"Row {i}: {val}\n")

print("Output written to comparison_output.txt")
