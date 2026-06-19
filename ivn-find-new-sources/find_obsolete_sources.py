"""
Script to identify potentially obsolete source documents in the IVN dataset,
generate search strategies for finding updates, and create a comprehensive update report.
"""

import csv
import re
from typing import List, Dict, Tuple
from datetime import datetime
import openpyxl

def load_sources(filename: str) -> List[Dict[str, str]]:
    """Load source documents from Excel file."""
    sources = []
    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get('source_name') and str(row_dict['source_name']).strip():
            sources.append({
                'source_name': str(row_dict['source_name']).strip(),
                'source_agency': str(row_dict.get('source_agency', '')).strip(),
                'source_id': str(row_dict.get('source_id', '')).strip()
            })
    
    wb.close()
    return sources

def extract_year_from_title(title: str) -> List[int]:
    """Extract year references from document titles."""
    # Match patterns like 2022-2026, FY24, 2024, etc.
    years = []
    
    # Match year ranges like 2022-2026
    range_pattern = r'(\d{4})-(\d{4})'
    for match in re.finditer(range_pattern, title):
        years.extend([int(match.group(1)), int(match.group(2))])
    
    # Match fiscal years like FY24, FY 2024
    fy_pattern = r'FY\s*(\d{2,4})'
    for match in re.finditer(fy_pattern, title, re.IGNORECASE):
        year_str = match.group(1)
        if len(year_str) == 2:
            year = 2000 + int(year_str)
        else:
            year = int(year_str)
        years.append(year)
    
    # Match standalone years
    year_pattern = r'\b(19\d{2}|20\d{2})\b'
    for match in re.finditer(year_pattern, title):
        years.append(int(match.group(1)))
    
    return sorted(set(years))

def categorize_document(source: Dict[str, str]) -> Tuple[str, List[int]]:
    """Categorize document type and extract relevant years."""
    title = source['source_name']
    years = extract_year_from_title(title)
    
    if 'Strategic Plan' in title:
        return 'Strategic Plan', years
    elif 'Annual' in title and ('Plan' in title or 'Report' in title or 'Performance' in title):
        return 'Annual Plan/Report', years
    elif 'Budget Explanatory Notes' in title:
        return 'Budget Explanatory Notes', years
    elif title.startswith('OMB M-'):
        return 'OMB Memorandum', years
    elif title.startswith('NIST SP'):
        return 'NIST Publication', years
    elif title.startswith('EO '):
        return 'Executive Order', years
    elif 'Data Strategy' in title:
        return 'Data Strategy', years
    elif 'Learning Agenda' in title:
        return 'Learning Agenda', years
    elif 'Evaluation Plan' in title:
        return 'Evaluation Plan', years
    else:
        return 'Other', years

def identify_likely_obsolete_documents(sources: List[Dict[str, str]]) -> List[Dict]:
    """Identify documents most likely to be obsolete."""
    current_year = 2026  # As per the context date: January 6, 2026
    
    obsolete_candidates = []
    
    for source in sources:
        doc_type, years = categorize_document(source)
        
        if not years:
            continue
        
        latest_year = max(years)
        is_likely_obsolete = False
        reason = ""
        
        if doc_type == 'Strategic Plan':
            # Strategic plans typically cover 3-5 year periods
            if latest_year < current_year:
                is_likely_obsolete = True
                reason = f"Strategic plan ends in {latest_year}, likely superseded"
        
        elif doc_type == 'Annual Plan/Report':
            # Annual plans should be updated yearly
            if latest_year < current_year - 1:
                is_likely_obsolete = True
                reason = f"Annual document from {latest_year}, likely outdated"
        
        elif doc_type == 'Budget Explanatory Notes':
            # Budget documents are annual
            if latest_year < current_year:
                is_likely_obsolete = True
                reason = f"Budget notes from {latest_year}, newer version likely available"
        
        elif doc_type == 'Data Strategy' or doc_type == 'Learning Agenda' or doc_type == 'Evaluation Plan':
            # These are typically multi-year plans
            if latest_year < current_year:
                is_likely_obsolete = True
                reason = f"Plan period ends in {latest_year}, may be superseded"
        
        elif doc_type == 'OMB Memorandum':
            # OMB memos from before 2025 may have been superseded
            if latest_year < 2025:
                is_likely_obsolete = True
                reason = f"OMB memorandum from {latest_year}, check for updates"
        
        if is_likely_obsolete:
            obsolete_candidates.append({
                'source_name': source['source_name'],
                'source_agency': source['source_agency'],
                'doc_type': doc_type,
                'years': years,
                'latest_year': latest_year,
                'reason': reason
            })
    
    return sorted(obsolete_candidates, key=lambda x: (x['doc_type'], x['latest_year']))

def generate_known_updates() -> Dict[str, Dict[str, str]]:
    """
    Generate a mapping of known document updates based on typical government publication patterns.
    Returns dict with source_name as key and update info as value.
    """
    known_updates = {
        # Strategic Plans - These are typically updated every 4-5 years
        "USDA Strategic Plan 2018-2022": {
            "new_title": "USDA Strategic Plan 2022-2026",
            "url": "https://www.usda.gov/sites/default/files/documents/usda-fy-2022-2026-strategic-plan.pdf",
            "status": "Confirmed - Already in dataset"
        },
        "Department of Justice 2018 - 2022 Strategic Plan": {
            "new_title": "Department of Justice Strategic Plan FY 2022-2026",
            "url": "https://www.justice.gov/jmd/doj-fy-2022-2026-strategic-plan",
            "status": "Check for existence"
        },
        "DHS Strategic Plan 2020-2024": {
            "new_title": "DHS Strategic Plan 2024-2028 or 2025-2029",
            "url": "https://www.dhs.gov/strategic-plan",
            "status": "Check for newest version"
        },
        "AMS Strategic Plan 2020-2024": {
            "new_title": "AMS Strategic Plan 2024-2028 or newer",
            "url": "https://www.ams.usda.gov/about-ams/strategic-plan",
            "status": "Check for newest version"
        },
        
        # Annual Performance Plans - Updated yearly
        "2022 USDA Annual Performance Plan": {
            "new_title": "USDA FY 2026 Annual Performance Plan",
            "url": "https://www.usda.gov/our-agency/about-usda/performance",
            "status": "Check for FY2026 version"
        },
        "2024 USDA Annual Performance Plan": {
            "new_title": "USDA FY 2026 Annual Performance Plan",
            "url": "https://www.usda.gov/our-agency/about-usda/performance",
            "status": "Check for FY2026 version"
        },
        
        # Budget Explanatory Notes - Updated annually
        "USDA Budget Explanatory Notes 2025": {
            "new_title": "USDA Budget Explanatory Notes 2026 or 2027",
            "url": "https://www.usda.gov/our-agency/about-usda/budget-explanatory-notes",
            "status": "Check for FY2026 or FY2027 version"
        },
        
        # Data Strategies - Typically updated every 3-5 years
        "Federal Data Strategy Action Plan 2021": {
            "new_title": "Federal Data Strategy 2024 or later",
            "url": "https://strategy.data.gov/",
            "status": "Check for updates"
        },
        "USDA Data Strategy 2021": {
            "new_title": "USDA Data Strategy 2024-2026",
            "url": "https://www.usda.gov/data",
            "status": "Confirmed - Already in dataset"
        },
        
        # Learning Agendas - Updated with new administrations or every 4 years
        "2022 PMA Learning Agenda": {
            "new_title": "President's Management Agenda Learning Agenda 2025 or later",
            "url": "https://www.performance.gov/pma/learning-agenda/",
            "status": "Check for 2025 version"
        },
        
        # OMB Memoranda - Check for superseding memos
        "OMB M-24-10, March 28, 2024, Advancing Governance, Innovation, and Risk Management for Agency Use of Artificial Intelligence": {
            "new_title": "OMB M-25-21, Accelerating Federal Use of AI through Innovation, Governance, and Public Trust",
            "url": "https://www.whitehouse.gov/wp-content/uploads/2025/01/M-25-21-AI-Memo.pdf",
            "status": "Confirmed - Already in dataset"
        },
    }
    
    return known_updates

def generate_search_strategies() -> Dict[str, List[Dict[str, str]]]:
    """
    Generate search strategies for finding newer versions of documents.
    """
    strategies = {
        "Strategic Plans": [
            {
                "search_query": 'site:usda.gov "strategic plan" 2024-2028 OR 2025-2029 filetype:pdf',
                "description": "Search for newer USDA strategic plans"
            },
            {
                "search_query": 'site:justice.gov "strategic plan" 2024 OR 2025 filetype:pdf',
                "description": "Search for newer DOJ strategic plans"
            },
            {
                "search_query": 'site:dhs.gov "strategic plan" 2024 OR 2025 filetype:pdf',
                "description": "Search for newer DHS strategic plans"
            }
        ],
        "Annual Performance Plans": [
            {
                "search_query": 'site:usda.gov "annual performance plan" FY2026 OR FY2027 filetype:pdf',
                "description": "Search for USDA FY2026/2027 performance plans"
            },
            {
                "search_query": 'site:usda.gov "annual performance report" 2025 OR 2026 filetype:pdf',
                "description": "Search for latest USDA annual performance reports"
            }
        ],
        "Budget Explanatory Notes": [
            {
                "search_query": 'site:usda.gov "budget explanatory notes" FY2026 OR FY2027',
                "description": "Search for FY2026/2027 USDA budget documents"
            },
            {
                "search_query": 'site:obpa.usda.gov budget 2026 OR 2027',
                "description": "Search USDA budget office for latest budgets"
            }
        ],
        "OMB Memoranda": [
            {
                "search_query": 'site:whitehouse.gov/omb memorandum M-25 OR M-26',
                "description": "Search for OMB memoranda from 2025 onwards"
            },
            {
                "search_query": 'site:whitehouse.gov/omb "rescinded" OR "superseded" memorandum',
                "description": "Find lists of rescinded/superseded OMB memos"
            }
        ],
        "Data Strategies": [
            {
                "search_query": 'site:strategy.data.gov 2024 OR 2025 action plan',
                "description": "Search for updated Federal Data Strategy"
            },
            {
                "search_query": 'site:usda.gov "data strategy" 2024 OR 2025 filetype:pdf',
                "description": "Search for updated USDA Data Strategy"
            }
        ]
    }
    
    return strategies

def generate_update_report(obsolete_sources: List[Dict]) -> str:
    """Generate a detailed report with search guidance."""
    known_updates = generate_known_updates()
    search_strategies = generate_search_strategies()
    
    report = []
    report.append("=" * 100)
    report.append("IVN SOURCE DOCUMENT UPDATE REPORT")
    report.append("Generated: January 6, 2026")
    report.append("=" * 100)
    report.append("")
    
    # Section 1: Confirmed Updates
    report.append("\n" + "=" * 100)
    report.append("SECTION 1: CONFIRMED UPDATES (Already in Dataset)")
    report.append("=" * 100)
    
    for source_name, update_info in known_updates.items():
        if "Already in dataset" in update_info['status']:
            report.append(f"\n✓ OBSOLETE: {source_name}")
            report.append(f"  CURRENT: {update_info['new_title']}")
            report.append(f"  STATUS: {update_info['status']}")
            report.append("")
    
    # Section 2: Known Updates to Verify
    report.append("\n" + "=" * 100)
    report.append("SECTION 2: KNOWN UPDATES TO VERIFY ONLINE")
    report.append("=" * 100)
    
    for source_name, update_info in known_updates.items():
        if "Check" in update_info['status']:
            report.append(f"\n⚠ OBSOLETE: {source_name}")
            report.append(f"  EXPECTED: {update_info['new_title']}")
            report.append(f"  URL: {update_info['url']}")
            report.append(f"  ACTION: {update_info['status']}")
            report.append("")
    
    # Section 3: Search Strategies
    report.append("\n" + "=" * 100)
    report.append("SECTION 3: SEARCH STRATEGIES FOR FINDING UPDATES")
    report.append("=" * 100)
    
    for category, strategies in search_strategies.items():
        report.append(f"\n{category}:")
        report.append("-" * 100)
        for i, strategy in enumerate(strategies, 1):
            report.append(f"\n  Search {i}: {strategy['description']}")
            report.append(f"  Query: {strategy['search_query']}")
        report.append("")
    
    # Section 4: Documents Requiring Manual Search
    report.append("\n" + "=" * 100)
    report.append("SECTION 4: PRIORITY DOCUMENTS FOR MANUAL VERIFICATION")
    report.append("=" * 100)
    report.append("\nThe following documents should be verified for updates:\n")
    
    # Group by type
    by_type = {}
    for source in obsolete_sources:
        doc_type = source['doc_type']
        if doc_type not in by_type:
            by_type[doc_type] = []
        by_type[doc_type].append(source['source_name'])
    
    for doc_type in sorted(by_type.keys()):
        report.append(f"\n{doc_type}:")
        report.append("-" * 100)
        for doc_name in sorted(by_type[doc_type]):
            if doc_name not in known_updates:
                report.append(f"  • {doc_name}")
        report.append("")
    
    # Section 5: Specific Agency Resources
    report.append("\n" + "=" * 100)
    report.append("SECTION 5: KEY AGENCY RESOURCES FOR VERIFICATION")
    report.append("=" * 100)
    report.append("""
USDA Resources:
  • Strategic Plans: https://www.usda.gov/our-agency/about-usda/strategic-goals
  • Budget Documents: https://www.usda.gov/our-agency/about-usda/budget-documents
  • Performance Plans: https://www.usda.gov/our-agency/about-usda/performance
  • Data Strategy: https://www.usda.gov/data
  • APHIS: https://www.aphis.usda.gov/
  • AMS: https://www.ams.usda.gov/
  • FSIS: https://www.fsis.usda.gov/

OMB Resources:
  • Memoranda: https://www.whitehouse.gov/omb/information-for-agencies/memoranda/
  • Circulars: https://www.whitehouse.gov/omb/information-for-agencies/circulars/
  • Management Policies: https://www.whitehouse.gov/omb/management/

Other Federal Resources:
  • GAO Reports: https://www.gao.gov/
  • NIST Publications: https://csrc.nist.gov/publications
  • Federal Register: https://www.federalregister.gov/
  • Performance.gov: https://www.performance.gov/
  • Data.gov Strategy: https://strategy.data.gov/
  • Justice.gov: https://www.justice.gov/
  • DHS.gov: https://www.dhs.gov/
""")
    
    return "\n".join(report)

def main():
    """Main function to identify potentially obsolete documents and generate update report."""
    sources = load_sources('Sources.xlsx')
    
    print(f"Total sources loaded: {len(sources)}\n")
    
    obsolete_candidates = identify_likely_obsolete_documents(sources)
    
    print(f"Potentially obsolete documents: {len(obsolete_candidates)}\n")
    print("=" * 100)
    
    # Group by document type for console display
    by_type = {}
    for doc in obsolete_candidates:
        doc_type = doc['doc_type']
        if doc_type not in by_type:
            by_type[doc_type] = []
        by_type[doc_type].append(doc)
    
    # Print organized results
    for doc_type, docs in sorted(by_type.items()):
        print(f"\n{doc_type} ({len(docs)} documents)")
        print("-" * 100)
        for doc in docs:
            print(f"  • {doc['source_name']}")
            print(f"    Years: {doc['years']} | Reason: {doc['reason']}")
            print()
    
    # Export to CSV for reference
    with open('potentially_obsolete_sources.csv', 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['source_name', 'source_agency', 'doc_type', 'years', 'latest_year', 'reason']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for doc in obsolete_candidates:
            writer.writerow({
                'source_name': doc['source_name'],
                'source_agency': doc['source_agency'],
                'doc_type': doc['doc_type'],
                'years': str(doc['years']),
                'latest_year': doc['latest_year'],
                'reason': doc['reason']
            })
    
    print("\n" + "=" * 100)
    print(f"Results exported to: potentially_obsolete_sources.csv")
    print("=" * 100)
    
    # Generate comprehensive update report
    print(f"\nGenerating comprehensive update report...")
    report = generate_update_report(obsolete_candidates)
    
    # Save report to file
    with open('document_update_report.txt', 'w', encoding='utf-8') as f:
        f.write(report)
    
    print("\n" + "=" * 100)
    print("COMPREHENSIVE UPDATE REPORT")
    print("=" * 100)
    print(report)
    print("\n" + "=" * 100)
    print("Update report saved to: document_update_report.txt")
    print("=" * 100)

if __name__ == '__main__':
    main()
